"""
=========================================================
NIFTY100 Platform

Sprint 3 - Peer Analytics Engine

Responsibilities
----------------
1. Load financial ratios from SQLite
2. Merge company information
3. Merge peer group mapping
4. Compute peer percentile rankings
5. Generate peer scores
6. Export SQLite tables
7. Export Excel reports
8. Generate radar charts

Author : Saranya
=========================================================
"""

from __future__ import annotations

import logging
import sqlite3

from dataclasses import dataclass
from pathlib import Path
from typing import Dict
from typing import List
from typing import Optional

import matplotlib

matplotlib.use("Agg")      # Must be before pyplot

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from src.database.database import DatabaseManager


# ==========================================================
# LOGGER
# ==========================================================

logger = logging.getLogger(__name__)

if not logger.handlers:

    logging.basicConfig(

        level=logging.INFO,

        format="%(asctime)s | %(levelname)s | %(message)s",

    )


# ==========================================================
# CUSTOM EXCEPTIONS
# ==========================================================

class PeerAnalyticsException(Exception):
    """
    Base Exception
    """


class PeerValidationError(PeerAnalyticsException):
    """
    Validation Error
    """


class PeerDatabaseError(PeerAnalyticsException):
    """
    Database Error
    """


# ==========================================================
# DATACLASSES
# ==========================================================

@dataclass(frozen=True)
class PeerMetric:
    """
    Metadata for one peer comparison metric.
    """

    column: str

    display_name: str

    higher_is_better: bool = True

    weight: float = 1.0


# ==========================================================
# PEER METRICS
# ==========================================================

PEER_METRICS: List[PeerMetric] = [

    PeerMetric(
        "return_on_equity_pct",
        "ROE",
        True,
        20,
    ),

    PeerMetric(
        "return_on_capital_employed_pct",
        "ROCE",
        True,
        15,
    ),

    PeerMetric(
        "net_profit_margin_pct",
        "Net Margin",
        True,
        10,
    ),

    PeerMetric(
        "debt_to_equity",
        "Debt/Equity",
        False,
        10,
    ),

    PeerMetric(
        "interest_coverage",
        "Interest Coverage",
        True,
        10,
    ),

    PeerMetric(
        "asset_turnover",
        "Asset Turnover",
        True,
        10,
    ),

    PeerMetric(
        "free_cash_flow",
        "Free Cash Flow",
        True,
        10,
    ),

    PeerMetric(
        "revenue_cagr_5yr",
        "Revenue CAGR",
        True,
        5,
    ),

    PeerMetric(
        "pat_cagr_5yr",
        "PAT CAGR",
        True,
        5,
    ),

    PeerMetric(
        "eps_cagr_5yr",
        "EPS CAGR",
        True,
        5,
    ),

]


# ==========================================================
# PEER ANALYTICS ENGINE
# ==========================================================

class PeerAnalyticsEngine:
    """
    Sprint 3 Peer Analytics Engine.

    Pipeline

    SQLite
        ↓

    financial_ratios

        +

    companies

        +

    peer_groups.xlsx

        ↓

    Master DataFrame

        ↓

    Percentile Ranking

        ↓

    Peer Scores

        ↓

    SQLite

        ↓

    Excel

        ↓

    Radar Charts
    """

    REQUIRED_RATIO_COLUMNS = [

        "company_id",

        "year",

        "return_on_equity_pct",

        "return_on_capital_employed_pct",

        "net_profit_margin_pct",

        "debt_to_equity",

        "interest_coverage",

        "asset_turnover",

        "free_cash_flow",

        "revenue_cagr_5yr",

        "pat_cagr_5yr",

        "eps_cagr_5yr",

    ]

    REQUIRED_COMPANY_COLUMNS = [

        "company_id",

        "company_name",

        "sector",

    ]

    REQUIRED_PEER_COLUMNS = [

        "company_id",

        "peer_group_name",

        "is_benchmark",

    ]

    # ------------------------------------------------------

    def __init__(

        self,

        peer_group_path: str,

        database_path: Optional[str] = None,

        output_dir: str = "output",

    ):

        self.peer_group_path = Path(
            peer_group_path
        )

        self.output_dir = Path(
            output_dir
        )

        self.output_dir.mkdir(

            parents=True,

            exist_ok=True,

        )

        self.report_dir = (

            self.output_dir /

            "peer_reports"

        )

        self.report_dir.mkdir(

            parents=True,

            exist_ok=True,

        )

        self.chart_dir = (
            Path("reports")
            / "radar_charts"
        )

        self.chart_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        

        self.db = DatabaseManager()

        self.database_path = database_path

        self.financial_df: Optional[
            pd.DataFrame
        ] = None

        self.company_df: Optional[
            pd.DataFrame
        ] = None

        self.peer_df: Optional[
            pd.DataFrame
        ] = None

        self.master_df: Optional[
            pd.DataFrame
        ] = None

        self.peer_percentiles: Optional[
            pd.DataFrame
        ] = None

        self.master_df: Optional[
            pd.DataFrame
        ] = None

        logger.info(
            "=" * 60
        )

        logger.info(
            "Peer Analytics Engine Initialized"
        )

        logger.info(
            "=" * 60
        )

    # ==========================================================
# DATABASE
# ==========================================================

    def connect(self):
        """
        Connect to SQLite database.
        """

        logger.info(
            "Connecting to database..."
        )

        self.db.connect()

        logger.info(
            "Database connected."
        )

    # ------------------------------------------------------

    def close(self):
        """
        Close database connection.
        """

        self.db.close()

        logger.info(
            "Database connection closed."
        )


# ==========================================================
# LOAD DATASETS
# ==========================================================

    def load_financial_ratios(
        self,
    ) -> pd.DataFrame:
        """
        Load financial_ratios table.
        """

        logger.info(
            "Loading financial ratios..."
        )

        query = """
        SELECT *
        FROM financial_ratios
        """

        self.financial_df = pd.read_sql_query(
            query,
            self.db.connection,
        )

        logger.info(
            "Financial Ratios : %d rows",
            len(self.financial_df),
        )

        return self.financial_df

    # ------------------------------------------------------

    def load_companies(
        self,
    ) -> pd.DataFrame:
        """
        Load companies.
        """

        logger.info(
            "Loading companies..."
        )

        query = """
        SELECT
            id,
            company_name
        FROM companies
        """

        self.company_df = pd.read_sql_query(
            query,
            self.db.connection,
        )

        self.company_df.rename(

            columns={

                "id": "company_id",

            },

            inplace=True,

        )

        logger.info(
            "Companies : %d",
            len(self.company_df),
        )

        return self.company_df

    # ------------------------------------------------------

    def load_peer_groups(
        self,
    ) -> pd.DataFrame:
        """
        Load peer_groups.xlsx
        """

        logger.info(
            "Loading peer groups..."
        )

        self.peer_df = pd.read_excel(
            self.peer_group_path
        )

        logger.info(
            "Peer Groups : %d",
            len(self.peer_df),
        )

        return self.peer_df


# ==========================================================
# VALIDATION
# ==========================================================

    @staticmethod
    def _validate_columns(
        dataframe: pd.DataFrame,
        required: List[str],
        dataset: str,
    ):

        missing = [

            column

            for column in required

            if column not in dataframe.columns

        ]

        if missing:

            raise PeerValidationError(

                f"{dataset} missing columns : {missing}"

            )

    # ------------------------------------------------------

    def validate(self):
        """
        Validate loaded datasets.
        """

        logger.info(
            "Validating datasets..."
        )

        if self.financial_df is None:
            raise PeerValidationError(
                "Financial dataframe not loaded."
            )

        if self.company_df is None:
            raise PeerValidationError(
                "Company dataframe not loaded."
            )

        if self.peer_df is None:
            raise PeerValidationError(
                "Peer dataframe not loaded."
            )

        self._validate_columns(

            self.financial_df,

            self.REQUIRED_RATIO_COLUMNS,

            "financial_ratios",

        )

        self._validate_columns(

            self.company_df,

            [

                "company_id",

                "company_name",

            ],

            "companies",

        )

        self._validate_columns(

            self.peer_df,

            self.REQUIRED_PEER_COLUMNS,

            "peer_groups",

        )

        logger.info(
            "Validation successful."
        )


# ==========================================================
# MASTER DATAFRAME
# ==========================================================

    def build_master_dataframe(
        self,
    ) -> pd.DataFrame:
        """
        Merge

        financial_ratios

        +

        companies

        +

        peer_groups
        """

        logger.info(
            "Building master dataframe..."
        )

        if self.financial_df is None:

            raise PeerAnalyticsException(
                "Financial dataframe not loaded."
            )

        if self.company_df is None:

            raise PeerAnalyticsException(
                "Company dataframe not loaded."
            )

        if self.peer_df is None:

            raise PeerAnalyticsException(
                "Peer dataframe not loaded."
            )

        df = self.financial_df.copy()

        companies = self.company_df.copy()

        peers = self.peer_df.drop(

            columns=["id"],

            errors="ignore",

        )

        df = df.merge(

            companies,

            how="left",

            on="company_id",

        )

        df = df.merge(

            peers,

            how="left",

            on="company_id",

        )

        self.master_df = df

        logger.info(
            "Master dataframe : %s",
            df.shape,
        )

        return df


    # ==========================================================
    # PREPARE
    # ==========================================================

    def prepare(
        self,
    ) -> pd.DataFrame:
        """
        Complete preparation pipeline.
        """

        self.connect()

        self.load_financial_ratios()

        self.load_companies()

        self.load_peer_groups()

        self.validate()

        return self.build_master_dataframe()

    # ==========================================================
# DATA CLEANING
# ==========================================================

    def keep_latest_year(
        self,
    ) -> pd.DataFrame:
        """
        Keep only the latest financial year
        for every company.
        """

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Master dataframe not available."
            )

        logger.info(
            "Keeping latest financial year..."
        )

        df = self.master_df.copy()

        df = (

            df.sort_values("year")

              .groupby(
                  "company_id",
                  as_index=False,
              )

              .tail(1)

              .reset_index(drop=True)

        )

        self.master_df = df

        logger.info(
            "Latest year companies : %d",
            len(df),
        )

        return df

    # ------------------------------------------------------

    def remove_duplicates(
        self,
    ) -> pd.DataFrame:

        """
        Remove duplicate companies.
        """

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Master dataframe not available."
            )

        before = len(self.master_df)

        self.master_df = (

            self.master_df

            .drop_duplicates(
                subset="company_id"
            )

            .reset_index(drop=True)

        )

        after = len(self.master_df)

        logger.info(
            "Duplicates removed : %d",
            before - after,
        )

        return self.master_df


# ==========================================================
# PERCENTILE CALCULATION
# ==========================================================

    @staticmethod
    def compute_metric_percentile(
        values: pd.Series,
        higher_is_better: bool,
    ) -> pd.Series:
        """
        Compute percentile rank.

        Higher values receive higher percentile.

        Debt ratio uses reverse ranking.
        """

        numeric = pd.to_numeric(
            values,
            errors="coerce",
        )

        if higher_is_better:

            return (
                numeric.rank(
                    pct=True,
                    method="average",
                ) * 100
            ).round(2)

        return (
            (
                1
                -
                numeric.rank(
                    pct=True,
                    method="average",
                )
            )
            * 100
        ).round(2)


# ==========================================================
# PEER PERCENTILES
# ==========================================================

    def compute_peer_percentiles(
        self,
    ) -> pd.DataFrame:
        """
        Compute percentile columns for every peer group.

        Percentiles are appended directly to
        master_df.
        """

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Master dataframe not available."
            )

        logger.info(
            "Computing peer percentiles..."
        )

        result = self.master_df.copy()

        grouped = result.groupby(
            "peer_group_name",
            dropna=False,
        )

        updated_groups = []

        for _, group in grouped:

            temp = group.copy()

            for metric in PEER_METRICS:

                if metric.column not in temp.columns:

                    continue

                percentile_column = (
                    f"{metric.column}_percentile"
                )

                temp[percentile_column] = (

                    self.compute_metric_percentile(

                        temp[metric.column],

                        metric.higher_is_better,

                    )

                )

            updated_groups.append(temp)

        result = pd.concat(
            updated_groups,
            ignore_index=True,
        )

        self.master_df = result

        logger.info(
            "Percentiles generated."
        )

        return result
# ==========================================================
# PEER SCORE
# ==========================================================

    def compute_peer_scores(
        self,
    ) -> pd.DataFrame:
        """
        Calculate weighted peer score.
        """

       
        df = self.master_df.copy()

        percentile_columns = [

            f"{metric.column}_percentile"

            for metric in PEER_METRICS

            if (
                f"{metric.column}_percentile" in df.columns
                and df[f"{metric.column}_percentile"].notna().any()
            )

        ]

        if not percentile_columns:

            raise PeerAnalyticsException(
                "Peer percentiles not available."
            )

        weighted_score = pd.Series(0.0, index=df.index)
        weight_sum = pd.Series(0.0, index=df.index)

        for metric in PEER_METRICS:

            column = f"{metric.column}_percentile"

            if column not in df.columns:
                continue

            valid = df[column].notna()

            weighted_score.loc[valid] += (
                df.loc[valid, column] * metric.weight
            )

            weight_sum.loc[valid] += metric.weight

        df["peer_score"] = (
            weighted_score / weight_sum.replace(0, np.nan)
        ).round(2)
       
        print()
        print("Companies with NaN score")
        print(

            df.loc[
                df["peer_score"].isna(),
                [
                    "company_id",
                    "company_name",
                    "peer_group_name",
                ]
            ]

        )
        df["peer_rank"] = (
        df.groupby("peer_group_name")["peer_score"]
        .rank(
            ascending=False,
            method="dense",
            na_option="bottom",
        )
        .astype("Int64")
    )
            
        self.master_df = df
        logger.info(
            "Peer scores completed."
        )

        return df
    # ==========================================================
    # SUMMARY
    # ==========================================================

    def peer_summary(
        self,
    ) -> Dict:

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Peer scores not available."
            )

        return {

            "peer_groups":

                self.master_df[
                    "peer_group_name"
                ].nunique(),

            "companies":

                self.master_df[
                    "company_id"
                ].nunique(),

            "metrics":

                len(PEER_METRICS),

            "average_peer_score":

                round(

                    self.master_df[
                        "peer_score"
                    ].mean(),

                    2,

                ),

        }

# ==========================================================
# SQLITE EXPORT
# ==========================================================

    def export_sqlite(
        self,
    ):
        """
        Export peer analytics to SQLite.
        """

        if (self.master_df is None or"peer_score" not in self.master_df.columns):
            raise PeerAnalyticsException( "Peer scores not available." )

            raise PeerAnalyticsException(
                "Peer scores not available."
            )

        logger.info(
            "Exporting peer analytics..."
        )

        conn = self.db.connection

        self.master_df.to_sql(

            "peer_percentiles",

            conn,

            if_exists="replace",

            index=False,

        )

        logger.info(
            "SQLite export completed."
        )


# ==========================================================
# EXCEL EXPORT
# ==========================================================

    def export_excel(
        self,
        filename: str = "peer_comparison.xlsx",
    ):
        """
        Export peer comparison workbook.

        Sprint 3 Day 20 requirements:
        - Exactly one sheet per peer group
        - Company ID and company name
        - 20 financial metric columns
        - 10 peer percentile columns
        - Green / yellow / red percentile formatting
        - Benchmark company highlighted in amber
        - Peer median summary row
        """

        if (
            self.master_df is None
            or "peer_score" not in self.master_df.columns
        ):
            raise PeerAnalyticsException(
                "Peer scores not available."
            )

        from openpyxl.styles import (
            Alignment,
            Font,
            PatternFill,
        )
        from openpyxl.utils import get_column_letter

        output = self.output_dir / filename

        output.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        logger.info(
            "Creating Excel report..."
        )

        # ----------------------------------------------------------
        # Report columns
        # ----------------------------------------------------------

        identity_columns = [
            "company_id",
            "company_name",
        ]

        metric_columns = [
            "return_on_equity_pct",
            "return_on_capital_employed_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "return_on_assets_pct",
            "debt_to_equity",
            "interest_coverage",
            "net_debt",
            "asset_turnover",
            "free_cash_flow",
            "cfo_quality_score",
            "fcf_conversion",
            "capex_intensity",
            "revenue_cagr_3yr",
            "revenue_cagr_5yr",
            "pat_cagr_3yr",
            "pat_cagr_5yr",
            "eps_cagr_3yr",
            "eps_cagr_5yr",
            "composite_quality_score",
        ]

        percentile_columns = [
            f"{metric.column}_percentile"
            for metric in PEER_METRICS
        ]

        ranking_columns = [
            "peer_score",
            "peer_rank",
            "is_benchmark",
        ]

        requested_columns = (
            identity_columns
            + metric_columns
            + percentile_columns
            + ranking_columns
        )

        # Defensive check in case schema changes later.
        report_columns = [
            column
            for column in requested_columns
            if column in self.master_df.columns
        ]

        # ----------------------------------------------------------
        # Excel styles
        # ----------------------------------------------------------

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="C6EFCE",
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            fgColor="FFEB9C",
        )

        red_fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE",
        )

        benchmark_fill = PatternFill(
            fill_type="solid",
            fgColor="F4B183",
        )

        median_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        median_font = Font(
            bold=True,
        )

        # ----------------------------------------------------------
        # Workbook
        # ----------------------------------------------------------

        with pd.ExcelWriter(
            output,
            engine="openpyxl",
        ) as writer:

            peer_groups = sorted(
                self.master_df[
                    "peer_group_name"
                ]
                .dropna()
                .unique()
            )

            for group in peer_groups:

                group_df = (
                    self.master_df
                    .loc[
                        self.master_df[
                            "peer_group_name"
                        ] == group
                    ]
                    .sort_values(
                        "peer_rank"
                    )
                    .copy()
                )

                # Keep only business-facing report columns.
                sheet = group_df[
                    report_columns
                ].copy()

                sheet_name = str(group)[:31]

                sheet.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )

                worksheet = writer.sheets[
                    sheet_name
                ]

                # --------------------------------------------------
                # Header formatting
                # --------------------------------------------------

                for cell in worksheet[1]:

                    cell.fill = header_fill
                    cell.font = header_font

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )

                worksheet.freeze_panes = "A2"

                worksheet.auto_filter.ref = (
                    worksheet.dimensions
                )

                # --------------------------------------------------
                # Column lookup
                # --------------------------------------------------

                column_map = {
                    cell.value: cell.column
                    for cell in worksheet[1]
                }

                # --------------------------------------------------
                # Percentile colour coding
                # --------------------------------------------------

                for percentile_column in percentile_columns:

                    if percentile_column not in column_map:
                        continue

                    column_number = column_map[
                        percentile_column
                    ]

                    for row_number in range(
                        2,
                        worksheet.max_row + 1,
                    ):

                        cell = worksheet.cell(
                            row=row_number,
                            column=column_number,
                        )

                        value = cell.value

                        if value is None:
                            continue

                        try:
                            value = float(value)
                        except (TypeError, ValueError):
                            continue

                        if value >= 75:

                            cell.fill = green_fill

                        elif value <= 25:

                            cell.fill = red_fill

                        else:

                            cell.fill = yellow_fill

                # --------------------------------------------------
                # Benchmark row
                # --------------------------------------------------

                if "is_benchmark" in column_map:

                    benchmark_column = column_map[
                        "is_benchmark"
                    ]

                    for row_number in range(
                        2,
                        worksheet.max_row + 1,
                    ):

                        benchmark_value = worksheet.cell(
                            row=row_number,
                            column=benchmark_column,
                        ).value

                        if benchmark_value in (
                            True,
                            1,
                            "1",
                            "True",
                            "TRUE",
                            "Yes",
                            "YES",
                        ):

                            for column_number in range(
                                1,
                                worksheet.max_column + 1,
                            ):

                                worksheet.cell(
                                    row=row_number,
                                    column=column_number,
                                ).fill = benchmark_fill

                # --------------------------------------------------
                # Peer median summary row
                # --------------------------------------------------

                median_row_number = (
                    worksheet.max_row + 1
                )

                worksheet.cell(
                    row=median_row_number,
                    column=1,
                    value="Peer Median",
                )

                # Median for all numeric report columns.
                numeric_columns = sheet.select_dtypes(
                    include="number"
                ).columns

                for column in numeric_columns:

                    if column not in column_map:
                        continue

                    median_value = sheet[
                        column
                    ].median()

                    if pd.isna(median_value):
                        continue

                    worksheet.cell(
                        row=median_row_number,
                        column=column_map[column],
                        value=float(median_value),
                    )

                # Format complete median row.
                for column_number in range(
                    1,
                    worksheet.max_column + 1,
                ):

                    cell = worksheet.cell(
                        row=median_row_number,
                        column=column_number,
                    )

                    cell.fill = median_fill
                    cell.font = median_font

                # --------------------------------------------------
                # Width / alignment
                # --------------------------------------------------

                for column_number in range(
                    1,
                    worksheet.max_column + 1,
                ):

                    column_letter = get_column_letter(
                        column_number
                    )

                    max_length = 0

                    for cell in worksheet[
                        column_letter
                    ]:

                        if cell.value is None:
                            continue

                        max_length = max(
                            max_length,
                            len(str(cell.value)),
                        )

                    worksheet.column_dimensions[
                        column_letter
                    ].width = min(
                        max(max_length + 2, 12),
                        35,
                    )

                # Keep company names readable.
                if "company_name" in column_map:

                    company_name_letter = (
                        get_column_letter(
                            column_map[
                                "company_name"
                            ]
                        )
                    )

                    worksheet.column_dimensions[
                        company_name_letter
                    ].width = 32

        logger.info(
            "Excel exported."
        )
# ==========================================================
# SUMMARY
# ==========================================================

    def summary(
        self,
    ) -> Dict:

        if (self.master_df is None or "peer_score" not in self.master_df.columns):
            raise PeerAnalyticsException("Peer scores not available." )

        return {

            "companies":

                len(
                    self.master_df
                ),

            "peer_groups":

                self.master_df[
                    "peer_group_name"
                ].nunique(),

            "average_peer_score":

                round(

                    self.master_df[
                        "peer_score"
                    ].mean(),

                    2,

                ),

            "top_company":

                self.master_df

                .sort_values(
                    "peer_score",
                    ascending=False,
                )

                .iloc[0][
                    "company_name"
                ],

        }


# ==========================================================
# PIPELINE
# ==========================================================

    def run(
        self,
    ) -> pd.DataFrame:

        logger.info("=" * 60)
        logger.info("Peer Analytics Started")
        logger.info("=" * 60)

        try:

            self.prepare()

            self.keep_latest_year()

            self.remove_duplicates()

            self.compute_peer_percentiles()

            self.compute_peer_scores()

            self.export_sqlite()

            self.export_excel()

            logger.info(
                "Peer Analytics Completed."
            )

            return self.master_df

        finally:

            self.close()

    # ==========================================================
# RADAR CHART
# ==========================================================

    def generate_radar_chart(
        self,
        company_id: str,
    ) -> Path:
        """
        Generate radar chart comparing
        company percentile against peer average.
        """

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Run analytics first."
            )

        df = self.master_df

        company = df.loc[
            df["company_id"] == company_id
        ]

        if company.empty:

            raise PeerAnalyticsException(
                f"{company_id} not found."
            )

        company = company.iloc[0]

        metrics = [
            metric
            for metric in PEER_METRICS
            if f"{metric.column}_percentile"
            in df.columns
        ]

        labels = [
            metric.display_name
            for metric in metrics
        ]

        company_values = [
            company[
                f"{metric.column}_percentile"
            ]
            for metric in metrics
        ]

        peer = df.loc[
            df["peer_group_name"]
            ==
            company["peer_group_name"]
        ]

        peer_average = [

            peer[
                f"{metric.column}_percentile"
            ].mean()

            for metric in metrics

        ]

        angles = np.linspace(
            0,
            2 * np.pi,
            len(metrics),
            endpoint=False,
        )

        company_values += company_values[:1]

        peer_average += peer_average[:1]

        angles = np.concatenate(
            [angles, [angles[0]]]
        )

        fig = plt.figure(
            figsize=(8, 8)
        )

        ax = plt.subplot(
            111,
            polar=True,
        )

        ax.plot(
            angles,
            company_values,
            linewidth=2,
            label=company["company_name"],
        )

        ax.fill(
            angles,
            company_values,
            alpha=0.20,
        )

        ax.plot(
            angles,
            peer_average,
            linestyle="--",
            linewidth=2,
            label="Peer Average",
        )

        ax.set_xticks(
            angles[:-1]
        )

        ax.set_xticklabels(
            labels
        )

        ax.set_ylim(0, 100)

        ax.legend(
            loc="upper right"
        )

        output = (
            self.chart_dir
            / f"{company_id}_radar.png"
        )

        plt.savefig(
            output,
            dpi=200,
            bbox_inches="tight",
        )

        plt.close()

        logger.info(
            "Radar chart generated : %s",
            company_id,
        )

        return output

# ==========================================================
# ALL CHARTS
# ==========================================================

    def generate_all_radar_charts(
        self,
    ):

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Analytics not available."
            )

        logger.info(
            "Generating radar charts..."
        )

       
        for company in self.master_df[
            "company_id"
        ]:

            try:

                self.generate_radar_chart(
                    company
                )

            except Exception as ex:

                logger.warning(
                    "%s : %s",
                    company,
                    ex,
                )

# ==========================================================
# PREVIEW
# ==========================================================

    def preview(
        self,
        rows: int = 5,
    ) -> pd.DataFrame:

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Run analytics first."
            )

        return self.master_df.head(
            rows
        )

# ==========================================================
# DATASET SUMMARY
# ==========================================================

    def dataset_summary(
        self,
    ) -> Dict:

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Analytics not available."
            )

        return {

            "companies":

                self.master_df[
                    "company_id"
                ].nunique(),

            "peer_groups":

                self.master_df[
                    "peer_group_name"
                ].nunique(),

            "metrics":

                len(
                    PEER_METRICS
                ),

            "columns":

                len(
                    self.master_df.columns
                ),

            "rows":

                len(
                    self.master_df
                ),

        }

# ==========================================================
# TOP COMPANIES
# ==========================================================

    def top_companies(
        self,
        n: int = 10,
    ) -> pd.DataFrame:

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Analytics not available."
            )

        return (

            self.master_df

            .sort_values(
                "peer_score",
                ascending=False,
            )

            .head(n)

        )

# ==========================================================
# RESET
# ==========================================================

    def reset(
        self,
    ):

        self.financial_df = None

        self.company_df = None

        self.peer_df = None

        self.master_df = None

        logger.info(
            "Peer engine reset."
        )
    # ==========================================================
# DATABASE TRANSACTION
# ==========================================================

    def execute_transaction(self):
        """
        Persist peer analytics using a
        single database transaction.
        """

        if self.master_df is None:

            raise PeerAnalyticsException(
                "No analytics available."
            )

        logger.info(
            "Saving peer analytics..."
        )

        conn = self.db.connection

        try:

            conn.execute("BEGIN")

            self.master_df.to_sql(

                "peer_percentiles",

                conn,

                if_exists="replace",

                index=False,

            )

            conn.commit()

            logger.info(
                "Peer analytics committed."
            )

        except Exception:

            conn.rollback()

            logger.exception(
                "Peer analytics rolled back."
            )

            raise


# ==========================================================
# MISSING VALUE REPORT
# ==========================================================

    def missing_value_report(
        self,
    ) -> pd.DataFrame:

        if self.master_df is None:

            raise PeerAnalyticsException(
                "Analytics not available."
            )

        report = pd.DataFrame(

            {

                "column":

                    self.master_df.columns,

                "missing":

                    self.master_df
                    .isna()
                    .sum()
                    .values,

                "missing_pct":

                    (

                        self.master_df
                        .isna()
                        .mean()

                        * 100

                    ).round(2).values,

            }

        )

        return report.sort_values(

            "missing",

            ascending=False,

        )


# ==========================================================
# AVAILABLE METRICS
# ==========================================================

    @staticmethod
    def available_metrics():

        return [

            metric.column

            for metric in PEER_METRICS

        ]


# ==========================================================
# STATUS
# ==========================================================

    def status(self):

        return {

            "financial_loaded":

                self.financial_df is not None,

            "companies_loaded":

                self.company_df is not None,

            "peer_loaded":

                self.peer_df is not None,

            "master_ready":

                self.master_df is not None,

        }


# ==========================================================
# REPORT GENERATOR
# ==========================================================

    def generate_reports(self):
        """
        Generate all reports.
        """

        logger.info(
            "Generating reports..."
        )

        self.export_excel()

        self.generate_all_radar_charts()

        logger.info(
            "Reports completed."
        )


# ==========================================================
# FULL PIPELINE
# ==========================================================

    def run(
        self,
    ) -> pd.DataFrame:

        logger.info("=" * 70)

        logger.info(
            "Sprint 3 Peer Analytics Started"
        )

        logger.info("=" * 70)

        try:

            self.prepare()

            self.keep_latest_year()

            self.remove_duplicates()

            self.compute_peer_percentiles()

            self.compute_peer_scores()

            self.execute_transaction()

            self.generate_reports()

            logger.info(
                "=" * 70
            )

            logger.info(
                "Companies : %d",

                len(self.master_df),

            )

            logger.info(
                "Peer Groups : %d",

                self.master_df[
                    "peer_group_name"
                ].nunique(),

            )

            logger.info(
                "=" * 70
            )

            return self.master_df

        finally:

            self.close()


# ==========================================================
# STRING REPRESENTATION
# ==========================================================

    def __repr__(self):

        if self.master_df is None:

            companies = 0

        else:

            companies = len(
                self.master_df
            )

        return (

            "PeerAnalyticsEngine("

            f"companies={companies})"

        )


# ==========================================================
# MODULE EXPORTS
# ==========================================================

__all__ = [

    "PeerAnalyticsEngine",

    "PeerMetric",

    "PEER_METRICS",

]


# ==========================================================
# MAIN
# ==========================================================

if __name__ == "__main__":

    engine = PeerAnalyticsEngine(

        peer_group_path=(
            "supporting_datasets/"
            "peer_groups.xlsx"
        )

    )

    engine.run()
    